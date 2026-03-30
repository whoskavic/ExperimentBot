import discord
import pytz
from datetime import datetime, time as dtime, timedelta
from flask import Flask, jsonify
from threading import Thread
import json
import asyncio
import os
import openpyxl

with open("appsettings.json", "r") as f:
    config = json.load(f)

with open("user.json", "r") as f:
    user_map = json.load(f)

TOKEN = config["bot_token"]
CHANNEL_IDS = config["channel_ids"]
THREAD_IDS = config["thread_ids"]
DESTINATION_CHANNEL_ID = config["destination_channel"]
RECAP_HOUR = config.get("recap_hour", 21)
RECAP_MINUTE = config.get("recap_minute", 0)

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)

tz = pytz.timezone("Asia/Jakarta")

# ===== Flask =====
app = Flask(__name__)

@app.route("/")
def home():
    return open("index.html", encoding="utf-8").read()

@app.route("/recap", methods=["POST"])
def trigger_recap_api():
    future = asyncio.run_coroutine_threadsafe(do_recap_api(), client.loop)
    try:
        files = future.result(timeout=120)
        return jsonify({"status": "ok", "files": files})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

def run_web():
    app.run(host="0.0.0.0", port=8080)

def keep_alive():
    t = Thread(target=run_web)
    t.daemon = True
    t.start()


# ===== HELPER =====
def resolve_name(author):
    uid = str(author.id)
    return user_map.get(uid, author.display_name).upper()


# ===== RECAP GENERATOR =====
async def generate_recap(source):
    now = datetime.now(tz)
    today = now.date()

    opening = {}
    closing = {}

    start_of_day = tz.localize(datetime.combine(today, dtime.min))

    async for msg in source.history(after=start_of_day, limit=None):
        msg_time = msg.created_at.astimezone(tz)
        lines = msg.content.splitlines()
        if not lines:
            continue

        header = lines[0].strip().upper()
        user = resolve_name(msg.author)

        if ("OPENING" in header or "OPEN" in header) and msg_time.time() <= dtime(10, 59, 59):
            opening[user] = 1

        if ("CLOSING" in header or "CLOSE" in header) and dtime(10, 59, 59) <= msg_time.time() <= dtime(20, 59, 59):
            closing[user] = 1

    users = sorted(set(opening.keys()) | set(closing.keys()))

    source_name = source.name.replace(" ", "_")
    filename = f"recap_{today}_{source_name}.xlsx" 

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recap"
    ws.append(["date", "source", "user", "opening", "closing"])
    for u in users:
        ws.append([
            str(today),
            source.name,
            u,
            1 if opening.get(u) else 0,
            1 if closing.get(u) else 0,
        ])
    wb.save(filename)

    return filename, today


# ===== RUN ALL RECAPS =====
async def run_all_recaps():
    results = []

    for channel_id in CHANNEL_IDS:
        source = client.get_channel(channel_id)
        if source is None:
            print(f"Channel {channel_id} not found")
            continue
        filename, today = await generate_recap(source)
        results.append((filename, today, source.name))

    for thread_id in THREAD_IDS:
        source = client.get_channel(thread_id)
        if source is None:
            print(f"Thread {thread_id} not found")
            continue
        filename, today = await generate_recap(source)
        results.append((filename, today, source.name))

    return results


# ===== KIRIM RECAP KE DISCORD =====
async def send_recaps(dest, results):
    today = datetime.now(tz).date()

    # Buat all_channel.xlsx dari semua hasil
    all_filename = f"recap_{today}_all_channel.xlsx"
    wb_all = openpyxl.Workbook()
    ws_all = wb_all.active
    ws_all.title = "All Channels"
    ws_all.append(["date", "source", "user", "opening", "closing"])

    for filename, _, _ in results:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            ws_all.append(list(row))

    wb_all.save(all_filename)

    # Kirim individual lalu hapus
    sent_names = []
    for filename, _, source_name in results:
        await dest.send(
            f"📋 Daily Recap `{today}` — **#{source_name}**",
            file=discord.File(filename)
        )
        sent_names.append(filename)
        os.remove(filename)

    # Kirim all_channel lalu hapus
    await dest.send(
        f"📋 Daily Recap `{today}` — **All Channels**",
        file=discord.File(all_filename)
    )
    sent_names.append(all_filename)
    os.remove(all_filename)

    return sent_names

# ===== API HANDLER (dari tombol UI) =====
async def do_recap_api():
    dest = client.get_channel(DESTINATION_CHANNEL_ID)
    results = await run_all_recaps()

    if not results:
        return []

    return await send_recaps(dest, results)


# ===== AUTO RECAP SCHEDULER jam 21:00 =====
async def schedule_recap():
    await client.wait_until_ready()

    while not client.is_closed():
        now = datetime.now(tz)
        target = now.replace(hour=RECAP_HOUR, minute=RECAP_MINUTE, second=0, microsecond=0)

        if now >= target:
            target += timedelta(days=1)

        wait_seconds = (target - now).total_seconds()
        print(f"Auto recap dijadwalkan: {target.strftime('%Y-%m-%d %H:%M:%S')} WIB")

        await asyncio.sleep(wait_seconds)

        dest = client.get_channel(DESTINATION_CHANNEL_ID)
        if dest is None:
            print("Destination channel not found")
            continue

        results = await run_all_recaps()

        if not results:
            await dest.send("Tidak ada data recap hari ini.")
            continue

        await send_recaps(dest, results)


# ===== BOT READY =====
@client.event
async def on_ready():
    print(f"Logged in as {client.user}")
    client.loop.create_task(schedule_recap())


# ===== COMMAND !recap =====
@client.event
async def on_message(message):
    if message.author == client.user:
        return

    if message.content.lower() == "!recap":
        current_id = message.channel.id

        if current_id not in CHANNEL_IDS and current_id not in THREAD_IDS:
            await message.channel.send("Command `!recap` tidak bisa dijalankan di sini.")
            return

        await message.channel.send("Generating recap...")

        source = client.get_channel(current_id)
        filename, today = await generate_recap(source)

        await message.channel.send(
            f"Recap `{today}` — **#{source.name}**",
            file=discord.File(filename)
        )
        os.remove(filename)


# ===== START =====
keep_alive()
client.run(TOKEN)