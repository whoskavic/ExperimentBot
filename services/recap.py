import asyncio
import os
import discord
import openpyxl
from datetime import datetime, date, time as dtime, timedelta

from config import TZ, USER_MAP, CHANNEL_IDS, THREAD_IDS, DESTINATION_CHANNEL


# ── helpers ──────────────────────────────────────────────

def resolve_name(author) -> str:
    return USER_MAP.get(str(author.id), author.display_name).upper()


def _start_of_day(d: date) -> datetime:
    return TZ.localize(datetime.combine(d, dtime.min))


def _end_of_day(d: date) -> datetime:
    return TZ.localize(datetime.combine(d, dtime(23, 59, 59)))


# ── time boundaries ───────────────────────────────────────

_OPEN_LATE   = dtime(10, 0, 0)   # opening dianggap telat jika > 10:00
_CLOSE_START = dtime(10, 0, 0)   # closing hanya valid mulai jam 10:00
_CLOSE_LATE  = dtime(20, 0, 0)   # closing dianggap telat jika > 20:00


# ── single source, full date range (1 API call) ───────────

# daily structure: {date: {user: {"op": bool, "cl": bool, "op_late": bool, "cl_late": bool}}}
type DayData = dict[str, dict[str, bool]]

async def _collect_range(source, date_from: date, date_to: date) -> dict[date, DayData]:
    """
    Fetch semua pesan dari 1 source untuk seluruh range dalam 1 API call.
    Return: {date: {user: {op, cl, op_late, cl_late}}}
    """
    daily: dict[date, DayData] = {}

    after  = _start_of_day(date_from)
    before = _end_of_day(date_to)

    async for msg in source.history(after=after, before=before, limit=None):
        msg_time = msg.created_at.astimezone(TZ)
        msg_date = msg_time.date()
        t        = msg_time.time()

        if msg.author.bot:
            continue

        if str(msg.author.id) not in USER_MAP:
            continue

        lines = msg.content.splitlines()
        if not lines:
            continue

        header = lines[0].strip().upper()

        is_open  = "OPENING" in header or "OPEN" in header
        is_close = ("CLOSING" in header or "CLOSE" in header) and t >= _CLOSE_START

        if not is_open and not is_close:
            continue

        user = resolve_name(msg.author)

        if msg_date not in daily:
            daily[msg_date] = {}
        if user not in daily[msg_date]:
            daily[msg_date][user] = {"op": False, "cl": False, "op_late": False, "cl_late": False}

        entry = daily[msg_date][user]

        if is_open:
            if t > _OPEN_LATE:
                entry["op_late"] = True
            else:
                entry["op"] = True

        if is_close:
            if t > _CLOSE_LATE:
                entry["cl_late"] = True
            else:
                entry["cl"] = True

    return daily


# ── generate xlsx for 1 source, date range ───────────────

async def generate_recap(source, date_from: date, date_to: date) -> str:
    """
    Generate .xlsx recap untuk 1 source (channel/thread),
    dari date_from sampai date_to (inklusif).
    Return: filepath string.
    """
    daily = await _collect_range(source, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recap"
    ws.append(["date", "source", "user", "opening", "opening_late", "closing", "closing_late"])

    current = date_from
    while current <= date_to:
        day_data = daily.get(current, {})
        for u in sorted(day_data):
            e = day_data[u]
            ws.append([
                str(current),
                source.name,
                u,
                1 if e["op"] else 0,
                1 if e["op_late"] else 0,
                1 if e["cl"] else 0,
                1 if e["cl_late"] else 0,
            ])
        current += timedelta(days=1)

    source_name = source.name.replace(" ", "_")
    label = f"{date_from}" if date_from == date_to else f"{date_from}_to_{date_to}"
    filename = f"recap_{label}_{source_name}.xlsx"
    wb.save(filename)
    return filename


# ── run all sources, date range ───────────────────────────

async def run_all_recaps(client, date_from: date, date_to: date) -> list[tuple[str, object]]:
    """
    Return list of (filepath, source_channel) untuk semua channel + thread.
    Semua source diproses secara paralel.
    """
    all_ids = [(cid, "channel") for cid in CHANNEL_IDS] + \
              [(tid, "thread")  for tid in THREAD_IDS]

    sources = []
    for source_id, kind in all_ids:
        source = client.get_channel(source_id)
        if source is None:
            print(f"{kind.capitalize()} {source_id} not found")
            continue
        sources.append(source)

    if not sources:
        return []

    filepaths = await asyncio.gather(
        *[generate_recap(src, date_from, date_to) for src in sources]
    )

    return list(zip(filepaths, sources))


# ── build all_channel.xlsx ────────────────────────────────

def build_all_channel_xlsx(results: list[tuple[str, object]], date_from: date, date_to: date) -> str:
    wb_all = openpyxl.Workbook()
    ws_all = wb_all.active
    ws_all.title = "All Channels"
    ws_all.append(["date", "source", "user", "opening", "opening_late", "closing", "closing_late"])

    for filepath, _ in results:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            ws_all.append(list(row))

    label = f"{date_from}" if date_from == date_to else f"{date_from}_to_{date_to}"
    all_filename = f"recap_{label}_all_channel.xlsx"
    wb_all.save(all_filename)
    return all_filename


# ── send to Discord & cleanup ─────────────────────────────

async def send_recaps(dest, results: list[tuple[str, object]], date_from: date, date_to: date) -> list[str]:
    """Kirim semua file ke dest channel (manual/API), hapus file lokal setelah kirim."""
    label = f"`{date_from}`" if date_from == date_to else f"`{date_from}` s/d `{date_to}`"

    all_filename = build_all_channel_xlsx(results, date_from, date_to)

    sent: list[str] = []

    for filepath, source in results:
        await dest.send(
            f"Recap {label} — **#{source.name}**",
            file=discord.File(filepath)
        )
        sent.append(filepath)
        os.remove(filepath)

    await dest.send(
        f"Recap {label} — **All Channels**",
        file=discord.File(all_filename)
    )
    sent.append(all_filename)
    os.remove(all_filename)

    return sent


async def send_recaps_scheduled(dest, results: list[tuple[str, object]], date_from: date, date_to: date) -> None:
    """
    Kirim recap scheduler: tiap file ke source channel-nya masing-masing,
    all_channel dikirim ke dest (destination_channel).
    """
    label = f"`{date_from}`" if date_from == date_to else f"`{date_from}` s/d `{date_to}`"

    all_filename = build_all_channel_xlsx(results, date_from, date_to)

    for filepath, source in results:
        await source.send(
            f"Recap {label} — **#{source.name}**",
            file=discord.File(filepath)
        )
        os.remove(filepath)

    await dest.send(
        f"Recap {label} — **All Channels**",
        file=discord.File(all_filename)
    )
    os.remove(all_filename)